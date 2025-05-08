import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from tkinter import messagebox, filedialog


def set_fixed_column_width(ws):
    """Set a fixed character width for the first 6 columns"""
    fixed_character_width = 15
    num_columns_to_set = 6

    for i in range(1, num_columns_to_set + 1):
        column_letter = get_column_letter(i)
        ws.column_dimensions[column_letter].width = fixed_character_width


def highlight_vacation_and_set_width(file_path):
    """Highlight vacation/leave rows in orange (planned, actual, Difference columns)."""
    wb = load_workbook(file_path)
    ws = wb.active

    #Set width
    set_fixed_column_width(ws)

    # Define orange fill color
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

    # Get column positions
    columns = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}  # header row is row 1

    # Loop over rows and highlight cells for vacation/leave projects
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        project_value = str(row[columns['Project'] - 1].value).lower()
        if 'vacation' in project_value or 'leave' in project_value:
            for col in ['Planned hours', 'Actual hours', 'Difference', 'Project', 'Month', 'Employee']:
                col_idx = columns[col] - 1
                row[col_idx].fill = orange_fill

    wb.save(file_path)


def save_final_excel(merged_df):
    """Save the merged DataFrame to Excel."""
    output_file_path = filedialog.asksaveasfilename(title="Save Excel file as",
                                                    defaultextension=".xlsx",
                                                    filetypes=[("Excel files", "*.xlsx")],
                                                    initialfile="Schedule_Differences.xlsx"
                                                    )

    if not output_file_path:
        messagebox.showinfo("Cancelled", "Save cancelled by user.")
        exit()

    merged_df.to_excel(output_file_path, index=False, sheet_name="Schedule Differences")
    highlight_vacation_and_set_width(output_file_path)

    messagebox.showinfo("Saved", f"Excel file successfully saved as: \n{output_file_path}")