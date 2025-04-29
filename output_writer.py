import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def save_final_excel(merged_df, output_file_name):
    """Save the merged DataFrame to Excel."""
    merged_df.to_excel(output_file_name, index=False, sheet_name="Schedule Differences")
    highlight_vacation_rows(output_file_name)

def highlight_vacation_rows(file_path):
    """Highlight vacation/leave rows in orange (planned, actual, diff columns)."""
    wb = load_workbook(file_path)
    ws = wb.active

    # Define orange fill color
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

    # Get column positions
    columns = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}  # header row is row 1

    # Loop over rows and highlight cells for vacation/leave projects
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        project_value = str(row[columns['project'] - 1].value).lower()
        if 'vacation' in project_value or 'leave' in project_value:
            for col in ['planned_hours', 'actual_hours', 'diff', 'project', 'month', 'employee']:
                col_idx = columns[col] - 1
                row[col_idx].fill = orange_fill

    wb.save(file_path)