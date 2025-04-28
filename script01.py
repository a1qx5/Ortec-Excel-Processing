import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from tkinter import Tk, filedialog, simpledialog

#Open a file picker for capacity planning file
Tk().withdraw()
capacity_file = filedialog.askopenfilename(title="Select Capacity Planning Excel File:", filetypes=[("Excel Files", "*xlsx")])

#Open a file picker for actual hours file
actual_file = filedialog.askopenfilename(title="Select Actual Hours Excel File:", filetypes=[("Excel Files", "*xlsx")])

#Ask user for output file name
output_file_name = simpledialog.askstring("Save As", "Enter the output Excel file name: ")
if '.xlsx' not in output_file_name:
    output_file_name += ".xlsx"

capacity_sheets = pd.read_excel(capacity_file, sheet_name=None)

def process_capacity_sheet(df, month_name):
    """Returns one capacity sheet"""

    #column for names
    df.columns = [str(col).strip() for col in df.columns]

    #employee and project columns
    id_cols = [col for col in df.columns  if col in ['Resource', 'Project']]

    #week columns
    week_cols = [col for col in df.columns if 'Week' in col]

    #wide -> long
    long_df = df.melt(id_vars = id_cols, value_vars = week_cols,
                      var_name = 'week', value_name = 'planned_days')

    #drop rows with no data
    long_df.dropna(subset=['planned_days'], inplace=True)

    # Convert commas to dots for decimal values (e.g. "0,5" → "0.5")
    long_df['planned_days'] = long_df['planned_days'].astype(str).str.replace(',', '.', regex=False)

    # Ensure planned_days is numeric
    long_df['planned_days'] = pd.to_numeric(long_df['planned_days'], errors='coerce')
    long_df.dropna(subset=['planned_days'], inplace=True)

    #month tag
    long_df['month'] = month_name

    #convert into hours
    long_df['planned_hours'] = long_df['planned_days'] * 8

    result = long_df.groupby(['Resource', 'Project', 'month'], as_index=False)['planned_hours'].sum()
    result.rename(columns={'Resource': 'employee', 'Project': 'project'}, inplace=True)

    return result

all_months = []
#automatically process all monthly sheets
for sheet_name, df in capacity_sheets.items():
    if "Resource Requirements - " in sheet_name:
        #Extract month name
        month = sheet_name.split(" - ")[-1]
        month_tag = f"{month}"

        #process and store the result
        processed = process_capacity_sheet(df, month_tag)
        all_months.append(processed)

#combine all monthly data into a single DataFrame
planned_df = pd.concat(all_months, ignore_index = True)

#Actual hours sheet
actual_df = pd.read_excel(actual_file, sheet_name = "Hours", skiprows = 2)

#normalize columns
actual_df.columns = actual_df.columns.str.strip().str.lower()

#drop rows without employee or project
actual_df.dropna(subset=['employee', 'project'], inplace=True)

#ensure they are strings
actual_df['employee'] = actual_df['employee'].astype(str)
actual_df['project'] = actual_df['project'].astype(str)

#detect month columns
month_columns = [col for col in actual_df.columns if col not in ['employee', 'project', 'total']]

#melt into long format
actual_long = actual_df.melt(id_vars=['employee', 'project'],
                             value_vars=month_columns,
                             var_name='month', value_name='actual_hours')

#handling the month column
actual_long['month'] = actual_long['month'].str.extract(r'([A-Za-z]+)')
actual_long['month'] = actual_long['month'].str.capitalize()

#clean
actual_long['actual_hours'] = actual_long['actual_hours'].astype(str).str.replace(',', '.', regex=False)
actual_long['actual_hours'] = pd.to_numeric(actual_long['actual_hours'], errors='coerce')
actual_long.dropna(subset=['actual_hours'], inplace=True)

actual_long = actual_long.groupby(['employee', 'project', 'month'], as_index=False)['actual_hours'].sum()

# Ensure employee and project are strings in both DataFrames
planned_df['employee'] = planned_df['employee'].astype(str)
planned_df['project'] = planned_df['project'].astype(str)
actual_long['employee'] = actual_long['employee'].astype(str)
actual_long['project'] = actual_long['project'].astype(str)

#merged sheet
merged_df = pd.merge(planned_df, actual_long, on=['employee', 'project', 'month'], how='outer')

#fill missing with 0
merged_df['planned_hours'] = merged_df['planned_hours'].fillna(0)
merged_df['actual_hours'] = merged_df['actual_hours'].fillna(0)

#calculate diff
merged_df['diff'] = merged_df['actual_hours'] - merged_df['planned_hours']

#handle vacation/leave rows
vacation_mask = merged_df['project'].str.lower().str.contains('vacation|leave', na=False)

#set diff = 0 for vacation/leave
merged_df.loc[vacation_mask, 'diff'] = 0

#calculate total_diff per employee and month
merged_df['total_diff'] = merged_df.groupby(['employee', 'month'])['diff'].transform('sum')

#only show total_diff at the last occurence per employee-month
merged_df['total_diff_per_month'] = ''

#find the last occurrence per employee_month
last_indexes = merged_df.groupby(['employee', 'month']).tail(1).index

#fill total_diff only for the last row
merged_df.loc[last_indexes, 'total_diff_per_month'] = merged_df.loc[last_indexes, 'total_diff']

#drop the old total_diff
merged_df.drop(columns=['total_diff'], inplace=True)

#calculate total_diff_per_project
####merged_df['total_diff_per_project'] = merged_df.groupby(['employee', 'project'])['diff'].transform('sum')

#sort
merged_df.sort_values(by=['employee', 'month', 'project'], inplace=True)

#save excel
merged_df.to_excel(output_file_name, index=False)
print(f"Excel file with merged data saved as {output_file_name}")


#excel cell styles
wb = load_workbook(output_file_name)
ws=wb.active

#vacation rows are highlighted in orange
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type='solid')

columns = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    project_cell = row[columns['project'] - 1]
    if project_cell.value and str(project_cell.value).lower().find('vacation') != -1 or str(project_cell.value).lower().find('leave') != -1:
        row[columns['employee'] - 1].fill = orange_fill
        row[columns['project'] - 1].fill = orange_fill
        row[columns['month'] - 1].fill = orange_fill
        row[columns['planned_hours'] - 1].fill = orange_fill
        row[columns['actual_hours'] - 1].fill = orange_fill
        row[columns['diff'] - 1].fill = orange_fill

#save the workbook
wb.save(output_file_name)

